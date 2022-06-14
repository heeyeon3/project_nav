import resource
from flask_restful import Resource, reqparse, request
from datetime import datetime, timedelta
from flask import jsonify

import json
import cv2
from markupsafe import re
from sqlalchemy.sql.elements import Null
from werkzeug.exceptions import PreconditionRequired
from config.properties import *
from models.excel import ExcelModel
from resource.location import locationlist
from resource.project import project_Edit_button
from utils.jsonutil import AlchemyEncoder as jsonEncoder
from utils.fileutil import FileUtils
import logging
import time
from flask import send_from_directory
import requests # excel backup
from openpyxl    import load_workbook, Workbook
from openpyxl.styles    import PatternFill
import os
from models.monitoring import *
from models.station import stationModel
from models.location import locationModel
from models.user import UserModel
from flask_login import current_user



class monitoring(Resource):

    parse = reqparse.RequestParser()

    parse.add_argument('station_id', type=str)
    parse.add_argument('monitoring_date', type=str)
    parse.add_argument('lastday', type=str)
    parse.add_argument('graph_count', type=str)

    # def get (self):
    #     result_string = {}
    #     monitoring_date =              request.args.get('monitoring_date')
    #     result_string = [dict(r) for r in stationModel.monitoring_lo_st(monitoring_date)]
        

        

    #     return {'resultCode': '0', "resultString": "SUCCESS", "data":result_string}, 200


    def post(self):

        params = monitoring.parse.parse_args()

        
        result_string = {}
        station_id =              params['station_id']
        monitoring_date =         params['monitoring_date']
        lastday =                 params['lastday']
        graph_count =             params['graph_count']
        print(graph_count)

        print(lastday,monitoring_date, "11111" )
        
        stationid_list = [dict(r) for r in stationModel.station_list()]
        print(stationid_list)
        station_id_list=[]

        for i in range(len(stationid_list)):
            station_id_list.append(stationid_list[i]['station_id'])

        

        try:
            #sql injection
            if len(monitoring_date) !=10 or len(lastday) != 10:
                return {'resultCode': '100', "resultString": "FAIL"}, 400

            test = '[\'\"\=\!\#\$\%\^\&\*\(\)\_\+\<\>\/\\\[\]]'
            result_project = ''

            if monitoring_date is not None :
                result_project = re.findall(test, monitoring_date)
                if len(result_project) > 0 : 
                    return {'resultCode': '100', "resultString": "FAIL"}, 400

            if lastday is not None :
                result_project = re.findall(test, lastday)
                if len(result_project) > 0 : 
                    return {'resultCode': '100', "resultString": "FAIL"}, 400

            if station_id not in station_id_list:
                return {'resultCode': '100', "resultString": "FAIL"}, 400
            # int(project_id)
            if graph_count and len(graph_count) > 0 :
                int(graph_count)
        
        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 400


        param = (station_id, monitoring_date,lastday )
    

        result_string = [dict(r) for r in monitoringModel.find_by_id_data(param)]

        today = str(datetime.now())
        today = today[0:10]
        print( today)

        if(graph_count == '0'):
            print("123")

        if(len(result_string)==0 and graph_count=='0'):
            
            i = 0
            while i<10:
                monitoring_date = datetime.strptime(monitoring_date,'%Y-%m-%d')
                monitoring_date = monitoring_date - timedelta(days=1)
                lastday = datetime.strptime(lastday,'%Y-%m-%d')
                lastday = lastday - timedelta(days=1)

                # print(monitoring_date,lastday, "22222" )

                monitoring_date = str(monitoring_date)
                monitoring_date = monitoring_date[0:10]
                lastday = str(lastday)
                lastday = lastday[0:10]

                # print(monitoring_date,lastday, "3333" )
                
                param = (station_id, monitoring_date,lastday )
                i = i+1
                print(i)
                result_string = [dict(r) for r in monitoringModel.find_by_id_data(param)]
                if(len(result_string)!=0):
                    break

            # print("종료" , result_string)
        
        return {'resultCode': '0', "resultString": "SUCCESS", "data":result_string}, 200


# data management
class data_file(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('projecttype', type=str)


    def get(self):
        result_string = {}
        projecttype =              request.args.get('projecttype')
      
        if(projecttype[0] == 'p'):
            station_id = projecttype[1:]
            result_string = [dict(r) for r in monitoringModel.find_station_id_data(station_id)]
        elif(projecttype[0] == 'l'):
            location_id = projecttype[1:]
            result_string = [dict(r) for r in monitoringModel.find_location_id_data(location_id)]
        elif(projecttype[0] == 's'):
            station_seq = projecttype[1:]
            result_string = [dict(r) for r in monitoringModel.find_station_id_data(station_seq)]

        return {'resultCode': '0', "resultString": "SUCCESS", "data":result_string}, 200



class Excel(Resource):
    parse = reqparse.RequestParser()
    parse.add_argument('startdate', type=str)
    parse.add_argument('enddate', type=str)
    parse.add_argument('datasave', type=str)
    parse.add_argument('ratedate', type=str)
    parse.add_argument('udp', type=str)
    parse.add_argument('excel_id', type=str)

    
    def get(self):
                
        startdate =              request.args.get('startdate')
        enddate =                request.args.get('enddate')
        datasave =               request.args.get('datasave')
        ratedate =               request.args.get('ratedate')
        udp =                    request.args.get('udp')

        stationid_list = [dict(r) for r in stationModel.station_list()]
        print(stationid_list)
        station_id_list=[]

        for i in range(len(stationid_list)):
            station_id_list.append(stationid_list[i]['station_id'])

        test = '[\'\"\,\.\=\!\#\$\%\^\&\*\(\)\_\+\<\>\/\\\[\]]'
        result_project = ''

        if udp is not None :
            result_project = re.findall(test, udp)
            if len(result_project) > 0 : 
                return {'resultCode': '100', "resultString": "FAIL"}, 400

        if startdate is not None :
            result_project = re.findall(test, startdate)
            if len(result_project) > 0 : 
                return {'resultCode': '100', "resultString": "FAIL"}, 400

        if enddate is not None :
            result_project = re.findall(test, enddate)
            if len(result_project) > 0 : 
                return {'resultCode': '100', "resultString": "FAIL"}, 400
        
        if len(result_project) > 0 : 
            return {'resultCode': '100', "resultString": "FAIL"}, 400

        

        try:
            firstword = datasave[0:1]
            otherword = datasave[1:len(datasave)]
            
            #sql injection
            if len(startdate) !=10 or len(enddate) != 10:
                return {'resultCode': '100', "resultString": "FAIL"}, 400

            if firstword != 'p' and firstword != 'l' and firstword != 's' :
                return {'resultCode': '100', "resultString": "FAIL"}, 400

            if otherword and len(otherword) > 0 :
                int(otherword)

            if ratedate and len(ratedate) > 0 :
                int(ratedate)            
        
        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 400


        user_id = current_user.user_id

        excel_request = startdate + "," + enddate + "," + datasave + "," + ratedate + "," + udp
        create_date = datetime.now()
        modify_date = datetime.now()

        print(excel_request)
        excel_obj = ExcelModel(excel_request, "request", None,user_id, create_date, modify_date)
    
        excel_obj.save_to_db()

        data = [dict(r) for r in ExcelModel.find_by_user_id(user_id, excel_request)]


        
        

        return {'resultCode': '0', "resultString": "SUCCESS", "data":data}, 200


    def post(self):
      
        params = Excel.parse.parse_args()

        excel_id =              params['excel_id']

        try :
            if excel_id and len(excel_id) > 0 :
                int(excel_id)
        
        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 400
        
        excel_obj = json.dumps(ExcelModel.find_by_id(excel_id), cls=jsonEncoder)

        return {'resultCode': '0', "resultString": "SUCCESS", "data":excel_obj}, 200


     

    



class ExcelDown(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('startdate', type=str)
    parse.add_argument('enddate', type=str)
    parse.add_argument('datasave', type=str)
    parse.add_argument('ratedate', type=str)
    parse.add_argument('udp', type=str)

    def get(param):
        # print("q")
        # print(param)

        # startdate, enddate, datasave, ratedate, udp = param
        startdate =              request.args.get('startdate')
        enddate =                request.args.get('enddate')
        datasave =               request.args.get('datasave')
        ratedate =               request.args.get('ratedate')
        udp =                    request.args.get('udp')
      
    

        param = (startdate, enddate, datasave, ratedate, udp)
   
   
        udp_list = []

        for i in range(len(udp)//5):
            if udp[i*5:i*5+5] == 'DDC01':
                udp_list.append('65001')
            elif udp[i*5:i*5+5] == 'DDC02':
                udp_list.append('65003')
            elif udp[i*5:i*5+5] == 'DDC03':
                udp_list.append('65005')
  

        now = datetime.now()
        stationlist = []
        if datasave[0] =='p':
            project_obj = [dict(r) for r in stationModel.find_project_id(datasave[1:])]

            if project_obj:
                file_name = project_obj[0]['project_name']+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"
                for i in range(len(project_obj)):
                    stationlist.append(project_obj[i]['station_id'])
       
                   
        elif datasave[0] =='l':
              location_obj  = [dict(r) for r in stationModel.find_location_id(datasave[1:])]
              if location_obj:
                file_name = location_obj[0]['location_name']+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"
                for i in range(len(location_obj)):
                    stationlist.append(location_obj[i]['station_id'])
        elif datasave[0] =='s':
            
            station_obj  = stationModel.find_by_seq(datasave[1:])
            file_name = station_obj.station_name+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"
            stationlist.append(station_obj.station_id)

      
        
        book = Workbook()
        # sheet = book.active

        # sheet.title = "device"

        # now = datetime.now()

        # file_name = datasave[1:]+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"
        # totalidx = 0
        sheet_length = 0
        for i in range(len(stationlist)):

            for j in range(len(udp_list)):
             
                param = (stationlist[i], startdate, enddate, ratedate, udp_list[j])
                udp_name = ""
                if udp_list[j]=='65001':
                    udp_name = "_DDC01"
                elif udp_list[j] == '65003':
                    udp_name = "_DDC02"
                elif udp_list[j] == '65005':                    
                    udp_name = "_DDC03"

                station_name = stationModel.find_by_id(stationlist[i]).station_name
                
                settop_list = [dict(r) for r in monitoringModel.find_save_data(param)]
            
                if len(settop_list) > 0:
                    sheet = book.create_sheet(station_name + udp_name)
                    sheet_length += 1
                    sheet.cell(row=1, column=1).value = "체크"
                    sheet.cell(row=1, column=2).value = "일시"
                    sheet.cell(row=1, column=3).value = "E"
                    sheet.cell(row=1, column=4).value = "N"
                    sheet.cell(row=1, column=5).value = "U"
                    sheet.cell(row=1, column=6).value = "WGS84 X"
                    sheet.cell(row=1, column=7).value = "WGS84 Y"
                    sheet.cell(row=1, column=8).value = "WGS84 Z"

                    sheet.cell(row=1, column=1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    sheet.cell(row=1, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    sheet.cell(row=1, column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    sheet.cell(row=1, column=4).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    sheet.cell(row=1, column=5).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    sheet.cell(row=1, column=6).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    sheet.cell(row=1, column=7).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    sheet.cell(row=1, column=8).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                

                    for idx in range(len(settop_list)):


                        sheet.cell(row=idx+2, column=2).value = settop_list[idx]['monitoring_date']
                        sheet.cell(row=idx+2, column=3).value = settop_list[idx]['monitoring_E']
                        sheet.cell(row=idx+2, column=4).value = settop_list[idx]['monitoring_N']
                        sheet.cell(row=idx+2, column=5).value = settop_list[idx]['monitoring_U']
                        sheet.cell(row=idx+2, column=6).value = settop_list[idx]['monitoring_X']
                        sheet.cell(row=idx+2, column=7).value = settop_list[idx]['monitoring_Y']
                        sheet.cell(row=idx+2, column=8).value = settop_list[idx]['monitoring_Z']

                
        
        if(sheet_length != 0):
            book.remove_sheet(book.get_sheet_by_name('Sheet'))
                    
                 
        
        
        book.save(filename = excel_file + file_name)

        # 다운로드 처리

        dls = excel_url+file_name
        resp = requests.get(dls)
        output = open(excel_file+file_name, 'rb')
        send_from_directory(excel_file, file_name)
        filename = excel_url + file_name

        return { "url" : filename }

class stationmap(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('location_id', type=str)

    def get (self):
        
        # SQL injection
        location_id =              request.args.get('location_id')

        try:
            if location_id and len(location_id) > 0 :
                int(location_id)
        
        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 400



        stationlist = [dict(r) for r in stationModel.monitoring_lo_st_station_id(location_id)]

        stationmapinfo = []
        # for i in range(len(stationlist)):
        #     station_id = stationlist[i]['station_id']
          

        #     find_info = [dict(r) for r in monitoringModel.find_mapinfo(station_id)]
        #     # print(find_info)
        #     if find_info:
        #         stationmapinfo.append(find_info[0])
        # print("stationmapinfo", stationmapinfo)

        station_id = ""
        for i in range(len(stationlist)):
            if(len(station_id) == 0 ):
                station_id += "'"+stationlist[i]['station_id']+"'"
            else:
                station_id += ", '"+stationlist[i]['station_id']+"'"
   

        stationmapinfo = [dict(r) for r in monitoringModel.find_mapinfo(station_id)]

 

        return {'resultCode': '0', "resultString": "SUCCESS", "data":stationmapinfo}, 200










class stationmonitoring(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('station_id', type=str)
    parse.add_argument('startdate', type=str)
    parse.add_argument('enddate', type=str)
    parse.add_argument('ratedate', type=str)
    parse.add_argument('udp', type=str)


    def post(self):


        params = stationmonitoring.parse.parse_args()

        station_id =              params['station_id']
        startdate =               params['startdate']
        enddate =                 params['enddate']
        ratedate =                params['ratedate']
        udp =                     params['udp']

        udp_list = ""


        
        stationid_list = [dict(r) for r in stationModel.station_list()]
        print(stationid_list)
        station_id_list=[]

        for i in range(len(stationid_list)):
            station_id_list.append(stationid_list[i]['station_id'])

        test = '[\'\"\=\!\#\$\%\^\&\*\(\)\_\+\<\>\/\\\[\]]'
        result_project = ''

        if udp is not None :
            result_project = re.findall(test, udp)

        if len(result_project) > 0 : 
            return {'resultCode': '100', "resultString": "FAIL"}, 400

        try:
            #sql injection
            if len(startdate) !=10 or len(enddate) != 10:
                return {'resultCode': '100', "resultString": "FAIL"}, 400

            if station_id not in station_id_list:
                return {'resultCode': '100', "resultString": "FAIL"}, 400

            test = '[\'\"\=\!\#\$\%\^\&\*\(\)\_\+\<\>\/\\\[\]]'
            result_project = ''

            if startdate is not None :
                result_project = re.findall(test, startdate)
                if len(result_project) > 0 : 
                    return {'resultCode': '100', "resultString": "FAIL"}, 400

            if enddate is not None :
                result_project = re.findall(test, enddate)
                if len(result_project) > 0 : 
                    return {'resultCode': '100', "resultString": "FAIL"}, 400

            if ratedate and len(ratedate) > 0 :
                int(ratedate)

            
        
        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 400
            



        for i in range(len(udp) // 5):
            if(len(udp_list) == 0):
                if udp[i*5:i*5+5] == 'DDC01':
                    udp_list += "'65001'"
                elif udp[i*5:i*5+5] == 'DDC02':
                    udp_list += "'65003'"
                elif udp[i*5:i*5+5] == 'DDC03':
                    udp_list += "'65005'"
            else:
                if udp[i*5:i*5+5] == 'DDC01':
                    udp_list += ", '65001'"
                elif udp[i*5:i*5+5] == 'DDC02':
                    udp_list += ", '65003'"
                elif udp[i*5:i*5+5] == 'DDC03':
                    udp_list += ", '65005'"
                
   
        param = (station_id, startdate, enddate, ratedate, udp_list)

        resultString = [dict(r) for r in monitoringModel.find_save_data2(param)]
  


        

        return {'resultCode': '0', "resultString": "SUCCESS", "data":resultString}, 200




class date_tooltip_monitoring(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('startdate', type=str)
    parse.add_argument('enddate', type=str)
    parse.add_argument('datasave', type=str)
    parse.add_argument('ratedate', type=str)

    def get(self):

        

        datasave =               request.args.get('datasave')
        
        

        #sql injection
        try:
            firstword = datasave[0:1]
            otherword = datasave[1:len(datasave)]

            if firstword != 'p' and firstword != 'l' and firstword != 's' :
                print(firstword)
                return {'resultCode': '100', "resultString": "FAIL"}, 400

            if otherword and len(otherword) > 0 :
                int(otherword)

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 400
            
        # print(datetime.now())

        datetimenow = datetime.now()
        datetimenow_30 = datetimenow - timedelta(days=30)
        datetimenow_30 = str(datetimenow_30)

        # format = '%Y-%m-%d %H:%M:%S'
        # datetimenow = datetime.strptime(datetimenow,format)
        # print(datetimenow_30)
        # print("datasave[1:]",datasave)
 
        station_id = ""
        if datasave[0] =='p':
            project_obj = [dict(r) for r in stationModel.find_project_id(datasave[1:])]
            if project_obj:
                for i in range(len(project_obj)):
                    if(i == 0):
                        station_id += "'"+ project_obj[i]['station_id'] + "'"
                    else:
                        station_id += ", '"+ project_obj[i]['station_id'] + "'"
        elif datasave[0] =='l':
              location_obj  = [dict(r) for r in stationModel.find_location_id(datasave[1:])]
              if location_obj:
                for i in range(len(location_obj)):
                    
                    if(i == 0):
                        station_id += "'"+ location_obj[i]['station_id'] + "'"
                    else:
                        station_id += ", '"+ location_obj[i]['station_id'] + "'"
        elif datasave[0] =='s':
                # station_id += "'"+ datasave[1:] + "'"

                station_obj  = stationModel.find_by_seq(datasave[1:])
                station_id = "'"+ station_obj.station_id + "'"
      
        resultString = [dict(r) for r in monitoringModel.tooltip_data_monitoring(station_id)]
        # unique = list({ each['monitoring_date'] : each for each in resultString }.values())

        # print("unique", unique)
  

        return {'resultCode': '0', "resultString": "SUCCESS", "data":resultString}, 200
        

class date_tooltip_satellite(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('startdate', type=str)
    parse.add_argument('enddate', type=str)
    parse.add_argument('datasave', type=str)
    parse.add_argument('ratedate', type=str)

    def get(self):

        datasave =               request.args.get('datasave')
        
        

        #sql injection
        try:
            firstword = datasave[0:1]
            otherword = datasave[1:len(datasave)]

            if firstword != 'p' and firstword != 'l' and firstword != 's' :
                print(firstword)
                return {'resultCode': '100', "resultString": "FAIL"}, 400

            if otherword and len(otherword) > 0 :
                int(otherword)

        except Exception as e:
            logging.fatal(e, exc_info=True)
            return {'resultCode': '100', "resultString": "FAIL"}, 400

 
        station_id = ""
        if datasave[0] =='p':
            project_obj = [dict(r) for r in stationModel.find_project_id(datasave[1:])]
            if project_obj:
                for i in range(len(project_obj)):
                    if(i == 0):
                        station_id += "'"+ project_obj[i]['station_id'] + "'"
                    else:
                        station_id += ", '"+ project_obj[i]['station_id'] + "'"
        elif datasave[0] =='l':
              location_obj  = [dict(r) for r in stationModel.find_location_id(datasave[1:])]
              if location_obj:
                for i in range(len(location_obj)):
                    
                    if(i == 0):
                        station_id += "'"+ location_obj[i]['station_id'] + "'"
                    else:
                        station_id += ", '"+ location_obj[i]['station_id'] + "'"
        elif datasave[0] =='s':
                # station_id += "'"+ datasave[1:] + "'"

                station_obj  = stationModel.find_by_seq(datasave[1:])
                station_id = "'"+ station_obj.station_id + "'"
       

        resultString = [dict(r) for r in monitoringModel.tooltip_data_satellite(station_id)]

 

        return {'resultCode': '0', "resultString": "SUCCESS", "data":resultString}, 200
   



class datastatus(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('status', type=str)

    def get(self):

        status = request.args.get('status')
        
        if status != "1000" :
            return {'resultCode': '100', "resultString": "FAIL"}, 400  

        station_OK = [dict(r) for r in monitoringModel.data_10min_status()]

        return {'resultCode': '0', "resultString": "SUCCESS", "data":station_OK}, 200

class datastatus_user(Resource):

    parse = reqparse.RequestParser()
    parse.add_argument('status', type=str)

    def get(self):

        status = request.args.get('status')
        
        if status != "1000" :
            return {'resultCode': '100', "resultString": "FAIL"}, 400  

        user_id =                       current_user.user_id
        user_obj =                      UserModel.find_by_id(user_id)
        
        user_gr =                  user_obj.user_gr
        print("CHECKHERE : ", user_id, user_gr)
        if(user_gr =='0101' or user_gr == '0102'):
            project_id = ""
            print("CHECKHERE IF : ", user_id, user_gr, project_id)

            station_OK = [dict(r) for r in monitoringModel.data_10min_status_user(project_id)]
        else:
            project_id =  user_obj.project_id
            print("CHECKHERE ELSE : ", user_id, user_gr, project_id)
            station_OK = [dict(r) for r in monitoringModel.data_10min_status_user(project_id)]
            


        return {'resultCode': '0', "resultString": "SUCCESS", "data":station_OK}, 200