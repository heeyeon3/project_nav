# import resource
# from flask_restful import Resource, reqparse, request
# from datetime import datetime, timedelta
# from flask import jsonify

# import json
# import cv2
# from sqlalchemy.sql.elements import Null
# from werkzeug.exceptions import PreconditionRequired
# from config.properties import *
# from resource.location import locationlist
# from resource.project import project_Edit_button
# from utils.jsonutil import AlchemyEncoder as jsonEncoder
# from utils.fileutil import FileUtils
# import logging
# import time
# from flask import send_from_directory
# import requests # excel backup
# from openpyxl    import load_workbook, Workbook
# from openpyxl.styles    import PatternFill
# import os
# from models.excel import ExcelModel

# from flask_login import current_user

# from posixpath import split
# import      socket
# import      mysql.connector

from datetime import datetime
import      pymysql
from openpyxl    import load_workbook, Workbook
from openpyxl.styles    import PatternFill
from config.properties import *
from flask_restful import Resource, reqparse, request
import requests # excel backup
from flask import send_from_directory
# import      mysql.connector



class exceldown:
    
    def exceldownload(self):
        print("CONNECT TRY")
        # mydb = pymysql.connect(
        #     host =         "218.38.232.231",
        #     user =          "navsys",
        #     password =      "navsysWkd!",
        #     database =      "NAVSYS",
        # # auth_plugin=    'mysql_native_password'
        # )

        mydb = pymysql.connect(
            host =         "115.68.110.145",
            user =          "navsys",
            password =      "navsysWkd!",
            database =      "NAVSYS",
        # auth_plugin=    'mysql_native_password'
        )
        try:

            mycursor = mydb.cursor()
            print("CONNECT mycursor")


            sql = "select * from tbl_excel where excel_status = 'request'"
            mycursor.execute(sql)
                
            # rowlist = mycursor.fetchone()
            rowlist = mycursor.fetchall()
            for row in rowlist:
                # print("111111")
                # print(row)
                excel_id = row[0]

                sql = "update tbl_excel set excel_status = 'download' where excel_id = '"+str(excel_id)+"';"
                mycursor.execute(sql)
                mydb.commit()

                excel_req = row[1]
                excel_req = excel_req.split(',')

                startdate = excel_req[0]
                enddate = excel_req[1]
                datasave = excel_req[2]
                ratedate = excel_req[3]
                udp = excel_req[4]

                excel_status = row[2]

                udp_list = []


                for i in range(len(udp)//5):
                    if udp[i*5:i*5+5] == 'DDC01':
                        udp_list.append('65001')
                    elif udp[i*5:i*5+5] == 'DDC02':
                        udp_list.append('65003')
                    elif udp[i*5:i*5+5] == 'DDC03':
                        udp_list.append('65005')

                now = datetime.now()
                stationlist = []
                stationnamelist = []


                if datasave[0] =='p':

                    sql = """select a.station_seq, a.station_id, a.station_name, c.location_id, c.location_name, d.project_id, d.project_name
                            from tbl_station a
                            left join tbl_location_station b on a.station_seq=b.station_seq
                            left join tbl_location c on b.location_id=c.location_id
                            left join tbl_project d on c.project_id = d.project_id
                            where a.use_yn = 'Y' and d.project_id='"""+ datasave[1:] +"';"

                    mycursor.execute(sql)
                    
                    project_obj = mycursor.fetchall()
                    if project_obj:
                        file_name = project_obj[0][6]+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"
                        for i in range(len(project_obj)):
                            stationlist.append(project_obj[i][1])
                            stationnamelist.append(project_obj[i][2])

                elif datasave[0] =='l':

                    sql = """select a.station_seq, a.station_id, a.station_name, c.location_id, c.location_name, d.project_id, d.project_name
                            from tbl_station a
                            left join tbl_location_station b on a.station_seq=b.station_seq
                            left join tbl_location c on b.location_id=c.location_id
                            left join tbl_project d on c.project_id = d.project_id
                            where a.use_yn = 'Y' and c.location_id='"""+ datasave[1:] +"';"

                    mycursor.execute(sql)
                    
                    location_obj = mycursor.fetchall()

                    if location_obj:
                        file_name = location_obj[0][6]+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"
                    for i in range(len(location_obj)):
                        stationlist.append(location_obj[i][1])
                        stationnamelist.append(location_obj[i][2])


                elif datasave[0] =='s':

                    sql = """select station_id, station_name from tbl_station where station_seq = '"""+datasave[1:] +"""';"""
                    mycursor.execute(sql)
                    
                    station_obj = mycursor.fetchall()
                    file_name = station_obj[0][1]+ now.strftime("%Y-%m-%d_%H:%M:%S")+".xlsx"
                    stationlist.append(station_obj[0][0])
                    stationnamelist.append(station_obj[0][1])



                book = Workbook()

                sheet_length = 0
                for i in range(len(stationlist)):

                    for j in range(len(udp_list)):
                        
                        param = (stationlist[i], startdate, enddate, ratedate, udp_list[j])
                        station_id = stationlist[i]
                        udpport = udp_list[j]
                        udp_name = ""
                        if udp_list[j]=='65001':
                            udp_name = "_DDC01"
                        elif udp_list[j] == '65003':
                            udp_name = "_DDC02"
                        elif udp_list[j] == '65005':                    
                            udp_name = "_DDC03"

                        # station_name = stationModel.find_by_id(stationlist[i]).station_name
                        station_name = stationnamelist[i]

                        # sql = """select station_id, date_format(monitoring_date, '%Y-%m-%d %H:%i:%S') monitoring_date, replace(FORMAT(monitoring_E,5), ',' , '') monitoring_E ,replace(FORMAT(monitoring_N,5), ',' , '') monitoring_N, 
                        #                     replace(FORMAT(monitoring_U,5), ',' , '') monitoring_U, replace(FORMAT(monitoring_X,5), ',' , '') monitoring_X,replace(FORMAT(monitoring_Y,5), ',' , '')  monitoring_Y, 
                        #                     replace(FORMAT(monitoring_Z,5), ',' , '') monitoring_Z, replace(FORMAT(monitoring_fix,5), ',' , '') monitoring_fix 
                        #                     from tbl_monitoring
                        #                     where station_id = '""" +station_id + "' and monitoring_date between date('"+startdate+"') and date('"+enddate+ """')+1 
                        #                     and monitoring_udp_port ='"""+udpport+"""' and mod(TIMESTAMPDIFF(MINUTE,(select monitoring_date from tbl_monitoring
                        #     where monitoring_date like '"""+startdate+"""%' and station_id = '""" +station_id + """'
                        #     limit 1),monitoring_date), """+ratedate+""")=0 and substr(monitoring_date, 18,2) =(select substr(monitoring_date, 18,2) from tbl_monitoring
                        #     where monitoring_date like '"""+startdate+"""%' and station_id = '""" +station_id + """'
                        #     limit 1)"""
                            


                        sql  = """select station_id, date_format(monitoring_date, '%Y-%m-%d %H:%i:%S') monitoring_date, replace(FORMAT(monitoring_E,5), ',' , '') monitoring_E ,replace(FORMAT(monitoring_N,5), ',' , '') monitoring_N, 
                                            replace(FORMAT(monitoring_U,5), ',' , '') monitoring_U, replace(FORMAT(monitoring_X,5), ',' , '') monitoring_X,replace(FORMAT(monitoring_Y,5), ',' , '')  monitoring_Y, 
                                            replace(FORMAT(monitoring_Z,5), ',' , '') monitoring_Z, replace(FORMAT(monitoring_fix,5), ',' , '') monitoring_fix
                                            from tbl_monitoring
                                            where station_id = '""" +station_id + "' and monitoring_date between date('"+startdate+"') and date('"+enddate+ """')+1 
                                            and monitoring_udp_port ='"""+udpport+"""' and mod(TIMESTAMPDIFF(MINUTE,(select monitoring_date from tbl_monitoring
                            where monitoring_date between date('"""+startdate+"""') and date('"""+enddate+ """')+1 and station_id = '""" +station_id + """'
                            limit 1),monitoring_date), """+ratedate+""")=0 and substr(monitoring_date, 18,2) =(select substr(monitoring_date, 18,2) from tbl_monitoring
                            where monitoring_date between date('"""+startdate+"""') and date('"""+enddate+ """')+1  and station_id = '""" +station_id + """'
                            limit 1)"""

                        mycursor.execute(sql)
                    
                        settop_list = mycursor.fetchall()
                        
                        # settop_list = [dict(r) for r in monitoringModel.find_save_data(param)]
                    
                        if len(settop_list) > 0:
                            sheet = book.create_sheet(station_name + udp_name)
                            sheet_length += 1
                            sheet.cell(row=1, column=1).value = "체크"
                            sheet.cell(row=1, column=2).value = "일시"
                            sheet.cell(row=1, column=3).value = "E"
                            sheet.cell(row=1, column=4).value = "N"
                            sheet.cell(row=1, column=5).value = "U"
                            sheet.cell(row=1, column=6).value = "WGS84 X"
                            sheet.cell(row=1, column=7).value = "WGS84 Y"
                            sheet.cell(row=1, column=8).value = "WGS84 Z"

                            sheet.cell(row=1, column=1).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet.cell(row=1, column=2).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet.cell(row=1, column=3).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet.cell(row=1, column=4).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet.cell(row=1, column=5).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet.cell(row=1, column=6).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet.cell(row=1, column=7).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            sheet.cell(row=1, column=8).fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                        

                            for idx in range(len(settop_list)):


                                sheet.cell(row=idx+2, column=2).value = settop_list[idx][1]
                                sheet.cell(row=idx+2, column=3).value = settop_list[idx][2]
                                sheet.cell(row=idx+2, column=4).value = settop_list[idx][3]
                                sheet.cell(row=idx+2, column=5).value = settop_list[idx][4]
                                sheet.cell(row=idx+2, column=6).value = settop_list[idx][5]
                                sheet.cell(row=idx+2, column=7).value = settop_list[idx][6]
                                sheet.cell(row=idx+2, column=8).value = settop_list[idx][7]

                        

                if(sheet_length != 0):
                    book.remove_sheet(book.get_sheet_by_name('Sheet'))


                book.save(filename = excel_file + file_name)

                # 다운로드 처리


                dls = excel_url+file_name
                resp = requests.get(dls)
                output = open(excel_file+file_name, 'rb')
                try:
                    send_from_directory(excel_file, file_name)
                except Exception as e:
                    print("************ excel DIRECTORY SAVE ERROR *************")
                    print(e)
                    
             
                filename = excel_url + file_name
                
                sql = """update tbl_excel set excel_status = 'complete' , excel_url ='"""+filename+"""' where excel_id = '"""+str(excel_id)+"""';"""
                mycursor.execute(sql)

                mydb.commit()

        except Exception as e:
            print("************ excel DOWN ERROR *************")
            print(e)
            # mydb = pymysql.connect(
            #     host =      "218.38.232.231",
            #     user =      "navsys",
            #     password =  "navsysWkd!",
            #     database =  "NAVSYS",
            #     # auth_plugin='mysql_native_password'
            # )

            mydb = pymysql.connect(
            host =         "115.68.110.145",
            user =          "navsys",
            password =      "navsysWkd!",
            database =      "NAVSYS",
        # auth_plugin=    'mysql_native_password'
        )
            # DB connection retry
            mycursor = mydb.cursor()

            mycursor.execute(sql)
            mydb.commit()

        mycursor.close()
        return
